import os

from openpyxl.reader.excel import load_workbook
from xmindparser import xmind_to_dict
import Template
from openpyxl import Workbook


#优先级映射关系
priority_mapping = {
    'priority-1': '高',
    'priority-2': '中',
    'priority-3': '低'
}

def load_xmind(file_path):
    '''加载xmind输出到字典项'''
    dict = xmind_to_dict(file_path)[0]['topic']
    return dict

def change_to_case(dict):
    '''
    将xmind获取的字典项按用例层级提取
    :param dict: xmind转换过来的字典
    :return: 全部用例
    '''
    # 分割每一条用例
    result = list()
    #遍历每个需求
    for story in dict:
        story_id = story['title']
        # 遍历每个模块
        for module in story['topics']:
            # 模块标题
            module_title = module['title']
            # 遍历每个用例
            for case in module['topics']:
                # 用例标题
                case_title = case['title']
                # 优先级
                if 'markers' in case:
                    priority = priority_mapping.get(case['markers'][0],' ')
                else:
                    priority = ' '
                # 遍历每个用例的前置、步骤和预期结果
                case_data = []
                for title in ['前置','步骤', '预期结果']:
                    for sub_topic in case['topics']:
                        if sub_topic['title'] == title:
                            case_data.append(sub_topic['topics'][0].get('title', ''))
                            break
                    else:
                        # 如果对应的子主题不存在，则插入空字符串
                        case_data.append('')

                #默认未执行
                case_data.append('NORUN')
                # 将用例数据添加到结果列表
                result.append([story_id,module_title, case_title,priority] + case_data)
    return result


def change_to_excel(xmind_path,excel_file):
    '''
    将读取到的用例写入excel
    '''
    dict = load_xmind(xmind_path)
    print(dict)
    #指定输出文件
    out_file=excel_file+'/'+dict.get('title','测试用例')+".xlsx"
    # 检查文件是否存在
    if os.path.exists(out_file):
        # 如果文件存在，打开现有工作簿
        wb = load_workbook(out_file)
        # 清空所有工作表
        for sheet in wb.sheetnames:
            wb.remove(wb[sheet])
    wb = Workbook()
    # 获取活动的工作表
    ws = wb.active
    # 给工作表命名
    ws.title = dict['title']
    # 写入标题行
    ws.append(['需求号','模块', '用例标题','优先级','前置条件', '步骤', '预期结果','实际结果'])
    #写入数据（将字典转换成单个用例列表）
    data=change_to_case(dict['topics'])
    for case in data:
        ws.append(case)

    #设置样式
    Template.set_style(ws)
    # 保存工作簿到文件
    wb.save(out_file)
    return out_file


if __name__=='__main__':
    change_to_excel('D:/Project/xmind转excel用例/用例模板.xmind','D:/Project/xmind转excel用例')