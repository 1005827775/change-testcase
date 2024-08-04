from openpyxl.styles import PatternFill, Border, Alignment, Side, Font, alignment


def set_style(ws):
    """
    设置表格模板
    """
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='d19a66', end_color='d19a66', fill_type='solid')  # 深蓝色背景
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
    header_alignment = Alignment(horizontal='center', vertical='center')

    # 设置表头样式
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    # 设置全表样式
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        for cell in col:
            cell.border = header_border

    # 设置列宽
    column_widths = {
        'A': 10,  # 需求号
        'B': 10,  # 模块
        'C': 40,  # 用例标题
        'D': 8,  # 优先级
        'E': 40,  # 前置条件
        'F': 40,  # 步骤
        'G': 40,  # 预期结果
        'H': 10  # 实际结果
        }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    #合并首列单元格
    merge_cells_by_first_column_value(ws,1)
    #合并第二列
    merge_cells_by_first_column_value(ws,2)


def merge_cells_by_first_column_value(ws,column,starting_row=2):
    """
    合并Excel工作表中首列相同值的单元格。

    :param ws: Excel工作表对象
    :param column: 合并的列
    :param starting_row: 开始合并的起始行，默认为第2行
    """
    try:
        # 加载第一列的值到列表中以减少对Excel文件的访问
        column_values = [ws.cell(row=row, column=column).value for row in range(starting_row, ws.max_row + 1)]
    except Exception as e:
        print(f"加载第一列值时发生错误: {e}")
        return

    end_row = starting_row + 1  # 初始化为起始行的前一行，以便下面逻辑加1后正好是起始行
    for row, value in enumerate(column_values, start=starting_row):
        try:
            # 检查当前行与下一行的值是否相同
            num= column_values[row-1]
            if value == num:
                end_row = end_row + 1  # 如果值相同，更新结束行
            else:
                # 值不同，进行合并，并重置结束行
                ws.merge_cells(start_row=starting_row, start_column=column, end_row=end_row-1, end_column=column)
                starting_row = end_row   # 更新起始行为当前行的下一行
                end_row = starting_row +1  # 重置结束行为起始行的前一行
        except IndexError:
            # 最后一行后面没有更多行，因此此错误可以安全忽略
            ws.merge_cells(start_row=starting_row, start_column=column, end_row=end_row-1, end_column=column)
        except Exception as e:
            print(f"处理行 {row} 时发生错误: {e}")
        # 设置居中对齐属性
        merged_cell = ws.cell(row=starting_row, column=column)
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')



